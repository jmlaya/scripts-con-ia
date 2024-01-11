import pymongo
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from bson import ObjectId, Timestamp

# Mapeo de tipos de Python a tipos de MongoDB
tipo_mongo = {
    'str': 'String',
    'int': 'Int32',
    'float': 'Double',
    'bool': 'Boolean',
    'ObjectId': 'ObjectId',
    'datetime': 'Date',
    'Timestamp': 'Timestamp',
    # Añade más mapeos según sea necesario
}

# Función para obtener el tipo de MongoDB
def get_mongo_type(val):
    if isinstance(val, ObjectId):
        return 'ObjectId'
    elif isinstance(val, Timestamp):
        return 'Timestamp'
    else:
        python_type = type(val).__name__
        return tipo_mongo.get(python_type, python_type)

# Función para analizar de manera recursiva campos anidados
def parse_nested_fields(document, prefix=''):
    fields = {}
    for key, value in document.items():
        value_type = get_mongo_type(value)

        if isinstance(value, (ObjectId, Timestamp)):
            value_example = str(value)
        elif isinstance(value, (dict, list)):
            value_example = str(value)
            value_type = 'String'
        else:
            value_example = value

        if isinstance(value, dict):
            nested_fields = parse_nested_fields(value, prefix=prefix + key + '.')
            fields.update(nested_fields)
        elif isinstance(value, list) and value and isinstance(value[0], dict):
            nested_fields = parse_nested_fields(value[0], prefix=prefix + key + '.')
            fields.update(nested_fields)
        else:
            fields[prefix + key] = {'type': value_type, 'example': value_example}
    return fields

# Función para obtener información de los índices de una colección
def get_indices_info(collection):
    indices = collection.index_information()
    indices_info = []
    for name, index in indices.items():
        index_fields = ', '.join([f'{k}: {v}' for k, v in index['key']])
        index_type = 'Unique' if index.get('unique', False) else 'Standard'
        indices_info.append({'name': name, 'type': index_type, 'fields': index_fields})
    return indices_info

# Función para aplicar bordes a todas las celdas de un rango
def set_border(ws, cell_range):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

# Conexión a MongoDB
client = pymongo.MongoClient("mongodb://127.0.0.1:27017")
db = client["dummy-db"]

# Crear un archivo de Excel
wb = Workbook()
wb.remove(wb.active)  # Remover la hoja por defecto si no se va a utilizar

collection_names = sorted(db.list_collection_names())

for collection_name in collection_names:
    collection = db[collection_name]
    
    # Analizar los documentos para obtener la estructura de los campos
    fields = {}
    for document in collection.find().limit(100):
        nested_fields = parse_nested_fields(document)
        for key, value in nested_fields.items():
            if key not in fields:
                fields[key] = value

    # Ordenar los campos alfabéticamente antes de crear el DataFrame
    sorted_fields = dict(sorted(fields.items()))

    # Si no hay campos, pasa a la siguiente colección
    if not sorted_fields:
        continue

    # Crear DataFrame para la colección
    df = pd.DataFrame.from_dict(sorted_fields, orient='index').reset_index()
    df.columns = ['field', 'type', 'example']

    # Obtener información de los índices
    indices_info = get_indices_info(collection)
    df_indices = pd.DataFrame(indices_info)

    # Agregar hoja al archivo de Excel
    sheet = wb.create_sheet(title=collection_name)
    
    # Estilos para los títulos de las secciones y los encabezados de las tablas
    section_title_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Título para la sección de campos
    sheet.append(['Fields'])
    fields_title_row = sheet.max_row
    fields_title_cell = sheet.cell(row=fields_title_row, column=1)
    fields_title_cell.font = section_title_font

    # Escribir la información de los campos
    fields_start_row = sheet.max_row + 1
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        sheet.append(r)
        if r_idx == 1:  # Aplicar estilo a la fila de encabezado
            for cell in sheet[sheet.max_row]:
                cell.fill = header_fill
                cell.font = header_font
    fields_end_row = sheet.max_row

    # Espacio entre las dos tablas
    sheet.append([])

    # Título para la sección de índices
    if not df_indices.empty:
        sheet.append(['Indexes'])
        indexes_title_row = sheet.max_row
        indexes_title_cell = sheet.cell(row=indexes_title_row, column=1)
        indexes_title_cell.font = section_title_font

        # Escribir la información de los índices
        indexes_start_row = sheet.max_row + 1
        for r_idx, r in enumerate(dataframe_to_rows(df_indices, index=False, header=True), 1):
            sheet.append(r)
            if r_idx == 1:  # Aplicar estilo a la fila de encabezado
                for cell in sheet[sheet.max_row]:
                    cell.fill = header_fill
                    cell.font = header_font
        indexes_end_row = sheet.max_row

        # Aplicar bordes a la tabla de índices
        if indexes_start_row <= indexes_end_row:
            set_border(sheet, f'A{indexes_start_row}:C{indexes_end_row}')

    # Aplicar bordes a la tabla de campos
    if fields_start_row <= fields_end_row:
        set_border(sheet, f'A{fields_start_row}:C{fields_end_row}')

    # Justificar a la izquierda todas las celdas de la columna 'C' excepto los encabezados
    for row in sheet.iter_rows(min_col=3, max_col=3, min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.row not in [fields_title_row + 1, indexes_title_row + 1]:
                cell.alignment = Alignment(horizontal='left')

# Guardar el archivo de Excel
wb.save('MongoDB_Documentation.xlsx')
